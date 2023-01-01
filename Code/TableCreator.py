import openpyxl, os
from openpyxl.styles import Font, Alignment

def update_table(planner, map, ext_mode, goal_bias, step_size, num_iter, time, cost, k=None, num_rewires=None):
    if not os.path.exists('Results.xlsx'):
        wb = openpyxl.Workbook()
        ws1 =  wb.active
        ws1.title = 'RRT'
        ws2 = wb.create_sheet('RRT Star')
        for idx, title in enumerate(['Map','Ext. Mode','Goal Bias','Step Size','Number of Iterations','Time [sec]','Cost']):
            ws1.cell(1,idx+1,value = title)
            ws1.cell(1,idx+1).font = Font(bold=True)
            ws1.cell(1,idx+1).alignment = Alignment(horizontal='center')
        for idx, title in enumerate(['Map','Ext. Mode','Goal Bias','Step Size','k','Number of Iterations','Time [sec]','Cost','Number of Rewirings']):
            ws2.cell(1,idx+1,value = title)
            ws2.cell(1,idx+1).font = Font(bold=True)
            ws2.cell(1,idx+1).alignment = Alignment(horizontal='center')

        ws1.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
        ws1.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 12
        ws1.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
        ws1.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 20
        ws1.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 12
        ws1.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
        
        ws2.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
        ws2.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 12
        ws2.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
        ws2.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 5
        ws2.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 20
        ws2.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 12
        ws2.column_dimensions[openpyxl.utils.get_column_letter(9)].width = 20
        
        wb.save(filename='Results.xlsx')
    else:
        wb = openpyxl.load_workbook('Results.xlsx')

    if planner=='rrt':
        ws = wb['RRT']
        i = 2
        while ws.cell(i,1).value is not None:
           i += 1
        ws.cell(i, 1, value = map)
        ws.cell(i, 2, value = ext_mode)
        ws.cell(i, 3, value = goal_bias)
        if ext_mode=='E2':
            ws.cell(i, 4, value = step_size)
        else:
            ws.cell(i, 4, value = '-')
        ws.cell(i, 5, value = num_iter)
        ws.cell(i, 6, value = f"{time:0.5f}")
        ws.cell(i, 7, value = f"{cost:0.3f}")
        for j in range(1,8):
            ws.cell(i,j).alignment = Alignment(horizontal='center')
    
    elif planner=='rrtstar':
        ws = wb['RRT Star']
        i = 2
        while ws.cell(i,1).value is not None:
           i += 1
        ws.cell(i, 1, value = map)
        ws.cell(i, 2, value = ext_mode)
        ws.cell(i, 3, value = goal_bias)
        if ext_mode=='E2':
            ws.cell(i, 4, value = step_size)
        else:
            ws.cell(i, 4, value = '-')
        ws.cell(i, 5, value = k)
        ws.cell(i, 6, value = num_iter)
        ws.cell(i, 7, value = f"{time:0.5f}")
        ws.cell(i, 8, value = f"{cost:0.3f}")
        ws.cell(i, 9, value = num_rewires)
        for j in range(1,10):
            ws.cell(i,j).alignment = Alignment(horizontal='center')
    
    wb.save(filename='Results.xlsx')
    return